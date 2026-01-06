import json
import os
from openpyxl import Workbook

from openpyxl.utils import get_column_letter

def convert_json_to_excel(input_json="data.json", output_excel="data.xlsx"):
    if not os.path.exists(input_json):
        print(f"Error: {input_json} not found.")
        return

    try:
        with open(input_json, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"Error reading JSON: {e}")
        return

    if not data or not isinstance(data, list):
        print("Error: JSON data is empty or not a list.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Dynamically determine headers from all records
    headers = []
    seen_headers = set()
    for row in data:
        if isinstance(row, dict):
            for key in row.keys():
                if key not in seen_headers:
                    headers.append(key)
                    seen_headers.add(key)

    if not headers:
        print("Error: No fields found in JSON objects.")
        return

    # Add header row
    ws.append(headers)

    # Add data rows
    for row in data:
        if isinstance(row, dict):
            ws.append([str(row.get(h, "")) for h in headers])

    # Auto-adjust column widths
    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(i)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50) # Cap at 50
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_excel)
    print(f"Successfully converted {input_json} to {output_excel}")

if __name__ == "__main__":
    convert_json_to_excel()
